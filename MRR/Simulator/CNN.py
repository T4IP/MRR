from re import A
import numpy as np
from typing import Any, Dict, List
import matplotlib.pyplot as plt
from tensorflow.keras.layers import Dense, Activation, Dropout, Conv1D, MaxPooling1D, GlobalAveragePooling1D
from tensorflow.keras.models import Sequential  
from tensorflow.keras.optimizers import Adam
import tensorflow as tf
import random
import os
import openpyxl
from tensorflow.python.keras.layers.core import Flatten
from transfer_function import simulate_transfer_function
from generate_figure import generate_figure

def minus(x):
    return -x

wb = openpyxl.load_workbook("C:\\Users\\taipa\\Documents\\研究室\\プログラム\\MRR\\LK.xlsx")    #エクセル読み込み
ws = wb.worksheets[0]   #ワークシート選択
data = []
for row in ws["I1:Q10"]:    #dataにLとKのデータセットを保存
    values = []
    for col in row:
        values.append(col.value)
    data.append(values)

number_of_rings = 8 #リング数
n = 10 #データ数
input_data = []
output_data = []
K = []
L=np.array([82.4e-6,82.4e-6,55.0e-6,82.4e-6,55.0e-6,82.4e-6,55.0e-6,55.0e-6])

for i in range(n):
    K = data[i][0:9]
    trans_data = simulate_transfer_function(L,K,config={'center_wavelength':1550e-9,'eta':0.996,'n_eff':2.2,'n_g':4.4,'alpha':53})
    input_data.append(list(map(minus,trans_data.simulate(np.arange(1540e-9,1560e-9,0.01e-9)))))
    output_data.append(K)

train_X = np.reshape(input_data,(-1,2001,1))
train_Y = np.reshape(output_data,(-1,9,1))
print(train_X.shape,train_Y.shape)

#シード値の固定
seed = 1
tf.random.set_seed(seed)
np.random.seed(seed)
random.seed(seed)
os.environ["PYTHONHASHSEED"] = str(seed)

#モデルの構造
epochs = 20
model_conv = Sequential()
model_conv.add(Conv1D(filters=32, kernel_size=3, strides=2, padding='same', input_shape=[2001,1], activation='linear'))
model_conv.add(Dropout(0.45))
model_conv.add(Conv1D(filters=32, kernel_size=3, strides=2, padding='same',activation='linear'))
model_conv.add(Dropout(0.30))
model_conv.add(Conv1D(filters=32, kernel_size=3, strides=2, padding='same',activation='linear'))
model_conv.add(MaxPooling1D(pool_size=2))
model_conv.add(Dropout(0.15))
#model_conv.add(GlobalAveragePooling1D())
model_conv.add(Flatten())
model_conv.add(Dense(9, activation='linear'))
model_conv.compile(optimizer = Adam(lr=0.001), loss='mean_squared_error')

#モデルの構造を表示する
print(model_conv.summary())

#予測開始
history = model_conv.fit(train_X, train_Y, batch_size=512,validation_split=0.1,epochs=epochs, verbose=1)

plt.plot(range(epochs),history.history["loss"],label="loss")
plt.plot(range(epochs),history.history["val_loss"],label="val_loss")
plt.xlabel("epoch")
plt.ylabel("loss")
plt.legend()
plt.show()

test = np.array(list(map(minus,np.array(generate_figure(1000,100,50,-30,2001)))))
pred_Y = model_conv.predict(test.reshape(1,2001,1))
print(pred_Y)

#model_conv.save("CNN.h5")

wb = openpyxl.load_workbook("C:\\Users\\taipa\\Documents\\研究室\\プログラム\\MRR\\LK2.xlsx")
ws = wb.worksheets[0]
test = []
for row in ws["A1:EXF10"]:
    values = []
    for col in row:
        values.append(col.value)
    test.append(values)
for i in range(10):
    pred = np.array(test[i][0:2001])
    pred_Y = model_conv.predict(np.array(pred.reshape(1,2001,1)))
    print(test[i][4001:4010])
    print(pred_Y)
